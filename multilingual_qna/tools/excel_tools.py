"""
Excel Output Tools for ADK Agents
==================================
Creates a professionally styled Excel workbook (QnA.xlsx) with
three language-specific sheets: English, Hindi, and Marathi.
"""

import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ─── Styling Constants ───────────────────────────────────────────────

SHEET_COLORS = {
    "English": "1F4E79",    # Deep Blue
    "Hindi": "C55A11",      # Warm Orange
    "Marathi": "548235",    # Forest Green
}

HEADER_FILLS = {
    "English": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    "Hindi": PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid"),
    "Marathi": PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
}

ROW_FILLS = {
    "even": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "odd": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

HEADER_BORDER = Border(
    left=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
    top=Side(style="thin", color="FFFFFF"),
    bottom=Side(style="medium", color="000000"),
)


def compile_excel_report(
    english_qna_json: list,
    hindi_qna_json: list,
    marathi_qna_json: list,
    output_path: str,
) -> str:
    """Compile QnA pairs from all three languages into a styled Excel file.

    Creates a professional Excel workbook with three color-coded sheets
    (English, Hindi, Marathi), each containing Questions and Answers columns
    with frozen headers, alternating row colors, and styled tab colors.

    Args:
        english_qna_json: List (or JSON string) of English QnA pairs array.
        hindi_qna_json: List (or JSON string) of Hindi QnA pairs array.
        marathi_qna_json: List (or JSON string) of Marathi QnA pairs array.
        output_path: File path to save the Excel file (e.g., 'QnA.xlsx').

    Returns:
        Confirmation message with file path and pair counts per language.
    """
    try:
        english_pairs = _parse_qna_json(english_qna_json, "English")
        hindi_pairs = _parse_qna_json(hindi_qna_json, "Hindi")
        marathi_pairs = _parse_qna_json(marathi_qna_json, "Marathi")
    except ValueError as e:
        return f"Error: {str(e)}"

    try:
        wb = Workbook()

        _create_sheet(wb, "English", english_pairs, is_first=True)
        _create_sheet(wb, "Hindi", hindi_pairs, is_first=False)
        _create_sheet(wb, "Marathi", marathi_pairs, is_first=False)

        # Ensure output directory exists
        out_dir = os.path.dirname(output_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)

        wb.save(output_path)

        return (
            f"SUCCESS: Excel file created at '{output_path}'.\n"
            f"- English sheet: {len(english_pairs)} QnA pairs\n"
            f"- Hindi sheet: {len(hindi_pairs)} QnA pairs\n"
            f"- Marathi sheet: {len(marathi_pairs)} QnA pairs\n"
            f"Total: {len(english_pairs) + len(hindi_pairs) + len(marathi_pairs)} pairs across 3 sheets."
        )
    except Exception as e:
        return f"Error creating Excel file: {str(e)}"


# ─── Internal Helpers ────────────────────────────────────────────────

def _parse_qna_json(qna_input, language: str) -> list:
    """Parse QnA input (which can be a list or a JSON string) into a list of dicts."""
    if isinstance(qna_input, list):
        return qna_input
    if isinstance(qna_input, dict):
        for key in ["qna", "qna_pairs", "pairs", "data"]:
            if key in qna_input and isinstance(qna_input[key], list):
                return qna_input[key]
        return [qna_input]
        
    try:
        cleaned = str(qna_input).strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        elif cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()

        # Try to find JSON array in the text
        json_match = re.search(r'\[.*\]', cleaned, re.DOTALL)
        if json_match:
            cleaned = json_match.group(0)

        pairs = json.loads(cleaned)
        if not isinstance(pairs, list):
            raise ValueError(f"{language}: Expected JSON array, got {type(pairs).__name__}")
        return pairs
    except json.JSONDecodeError as e:
        raise ValueError(f"{language}: Invalid JSON — {str(e)}")



def _create_sheet(wb: Workbook, sheet_name: str, qna_pairs: list, is_first: bool = False):
    """Create and style a single worksheet with QnA data."""
    if is_first:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    ws.sheet_properties.tabColor = SHEET_COLORS.get(sheet_name, "000000")

    # Header styling
    header_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = HEADER_FILLS.get(sheet_name, HEADER_FILLS["English"])

    # Write headers
    headers = ["Questions", "Answers"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = HEADER_BORDER

    # Data styling
    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Write QnA data rows
    for row_idx, pair in enumerate(qna_pairs, start=2):
        question = pair.get("question", pair.get("Question", ""))
        answer = pair.get("answer", pair.get("Answer", ""))

        q_cell = ws.cell(row=row_idx, column=1, value=question)
        a_cell = ws.cell(row=row_idx, column=2, value=answer)

        fill = ROW_FILLS["even"] if row_idx % 2 == 0 else ROW_FILLS["odd"]
        for cell in [q_cell, a_cell]:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = THIN_BORDER
            cell.fill = fill

    # Column widths and header row height
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 65
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
