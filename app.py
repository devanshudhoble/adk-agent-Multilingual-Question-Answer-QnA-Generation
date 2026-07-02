"""
Multilingual QnA Generator — Streamlit UI
============================================
Premium Streamlit interface for the ADK-powered Multilingual QnA Generation System.
Invokes the Google ADK multi-agent pipeline for document processing, QnA generation,
translation, and Excel export.

Usage:
    streamlit run app.py
"""

import streamlit as st
import asyncio
import json
import os
import re
import time
import tempfile
import shutil


# ─── Load API Key from .env ──────────────────────────────────────────
def _load_api_key() -> str:
    """Load Gemini API key from .env file or environment variable."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    if "GOOGLE_API_KEY" in line:
                        return line.split("=", 1)[1].strip().strip('"').strip("'")
    return os.environ.get("GOOGLE_API_KEY", "")


DEFAULT_API_KEY = _load_api_key()

# ─── Page Configuration ──────────────────────────────────────────────
st.set_page_config(
    page_title="Multilingual QnA Generator — ADK Agent System",
    page_icon="🌐",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Premium CSS ─────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    /* Gradient Header */
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 10px 40px rgba(102, 126, 234, 0.3);
    }
    .main-header h1 {
        color: white;
        font-size: 2.2rem;
        font-weight: 700;
        margin: 0;
        letter-spacing: -0.5px;
    }
    .main-header p {
        color: rgba(255,255,255,0.85);
        font-size: 1.05rem;
        margin-top: 0.5rem;
        font-weight: 300;
    }

    /* Feature badges */
    .feature-row {
        display: flex;
        justify-content: center;
        gap: 12px;
        margin-top: 1rem;
        flex-wrap: wrap;
    }
    .feature-badge {
        background: rgba(255,255,255,0.2);
        color: white;
        padding: 6px 16px;
        border-radius: 20px;
        font-size: 0.85rem;
        font-weight: 500;
        backdrop-filter: blur(10px);
        border: 1px solid rgba(255,255,255,0.15);
    }

    /* Agent Status Cards */
    .agent-card {
        background: linear-gradient(135deg, #f8f9fc 0%, #eef1f8 100%);
        border: 1px solid #e0e4ef;
        border-left: 4px solid #667eea;
        border-radius: 10px;
        padding: 1rem 1.2rem;
        margin-bottom: 0.6rem;
        transition: all 0.3s ease;
    }
    .agent-card.active {
        border-left-color: #00b894;
        background: linear-gradient(135deg, #f0fff4 0%, #e6ffed 100%);
        box-shadow: 0 4px 12px rgba(0, 184, 148, 0.15);
    }
    .agent-card.done {
        border-left-color: #00b894;
        opacity: 0.7;
    }
    .agent-name {
        font-weight: 600;
        font-size: 0.95rem;
        color: #1a1a2e;
    }
    .agent-desc {
        font-size: 0.8rem;
        color: #666;
        margin-top: 2px;
    }

    /* Step indicators */
    .step-indicator {
        display: flex;
        align-items: center;
        gap: 10px;
        padding: 12px 18px;
        background: #eef2ff;
        border-radius: 10px;
        margin-bottom: 0.8rem;
        border-left: 4px solid #667eea;
    }
    .step-number {
        background: #667eea;
        color: white;
        width: 28px;
        height: 28px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: 700;
        font-size: 0.85rem;
        flex-shrink: 0;
    }
    .step-text {
        color: #333;
        font-weight: 500;
        font-size: 0.95rem;
    }

    /* Stats row */
    .stats-row {
        display: flex;
        gap: 12px;
        margin: 1rem 0;
    }
    .stat-box {
        flex: 1;
        background: linear-gradient(135deg, #f5f7ff, #eef2ff);
        border: 1px solid #d4daef;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
    }
    .stat-value {
        font-size: 1.5rem;
        font-weight: 700;
        color: #667eea;
    }
    .stat-label {
        font-size: 0.8rem;
        color: #666;
        margin-top: 2px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    /* Success banner */
    .success-banner {
        background: linear-gradient(135deg, #00b894 0%, #00cec9 100%);
        color: white;
        padding: 1.2rem 1.5rem;
        border-radius: 12px;
        text-align: center;
        font-weight: 600;
        font-size: 1.1rem;
        margin: 1rem 0;
        box-shadow: 0 6px 20px rgba(0, 184, 148, 0.3);
    }

    /* Info card */
    .info-card {
        background: #f8f9fc;
        border: 1px solid #e8ecf4;
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1rem;
    }
    .info-card h3 {
        color: #1a1a2e;
        margin-top: 0;
        font-size: 1.1rem;
    }

    /* Architecture diagram */
    .arch-card {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border-radius: 14px;
        padding: 1.5rem 2rem;
        color: #e0e0e0;
        font-family: 'Courier New', monospace;
        font-size: 0.85rem;
        line-height: 1.6;
        margin: 1rem 0;
        box-shadow: 0 8px 30px rgba(0,0,0,0.2);
    }
    .arch-card .highlight {
        color: #667eea;
        font-weight: bold;
    }
    .arch-card .green {
        color: #00b894;
    }
    .arch-card .orange {
        color: #fdcb6e;
    }
    .arch-card .pink {
        color: #fd79a8;
    }

    /* Dark sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1a2e 0%, #16213e 100%);
    }
    [data-testid="stSidebar"] * {
        color: #e0e0e0 !important;
    }
    [data-testid="stSidebar"] .stMarkdown h1,
    [data-testid="stSidebar"] .stMarkdown h2,
    [data-testid="stSidebar"] .stMarkdown h3 {
        color: #ffffff !important;
    }

    /* Hide defaults */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ─── Sidebar ─────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Configuration")
    st.markdown("---")

    demo_mode = st.checkbox(
        "🤖 Demo Mode (Offline / No Key)",
        value=not bool(DEFAULT_API_KEY),
        help="Check this to test the app offline with mock data if you don't have an API key.",
    )

    if demo_mode:
        api_key = "DEMO_MODE"
        st.info("🤖 **Demo Mode Active**: Generates QnA pairs using ADK MockLlm offline.")
    else:
        api_key = st.text_input(
            "🔑 Google API Key",
            value=DEFAULT_API_KEY,
            type="password",
            help="Your Google AI Studio API key. Pre-loaded from .env file.",
        )

    st.markdown("---")

    num_pairs = st.slider(
        "📊 Number of QnA Pairs",
        min_value=5,
        max_value=25,
        value=10,
        step=1,
        help="Number of Question-Answer pairs to generate per language.",
    )

    model_name = st.selectbox(
        "🤖 Gemini Model",
        options=["gemini-2.0-flash", "gemini-1.5-flash", "gemini-2.5-flash"],
        index=0,
        help="Select the Gemini model for all agents.",
    )

    st.markdown("---")
    st.markdown("### 📁 Supported Formats")
    st.markdown("- 📄 **PDF** (.pdf)\n- 📝 **Word** (.docx)\n- 📃 **Text** (.txt)")
    st.markdown("---")
    st.markdown("### 🌐 Output Languages")
    st.markdown("- 🇬🇧 **English**\n- 🇮🇳 **Hindi** (हिन्दी)\n- 🇮🇳 **Marathi** (मराठी)")
    st.markdown("---")

    st.markdown("### 🤖 ADK Agent Pipeline")
    st.markdown(
        '<div class="arch-card">'
        '<span class="highlight">SequentialAgent</span>: Pipeline<br>'
        '├── <span class="green">📄 document_parser</span><br>'
        '├── <span class="orange">🧠 qna_generator</span><br>'
        '├── <span class="pink">🇮🇳 hindi_translator</span><br>'
        '├── <span class="pink">🇮🇳 marathi_translator</span><br>'
        '└── <span class="green">📊 excel_compiler</span>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        "<div style='text-align:center; opacity:0.5; font-size:0.8rem;'>"
        "Powered by Google ADK + Gemini API"
        "</div>",
        unsafe_allow_html=True,
    )


# ─── Main Content ────────────────────────────────────────────────────

# Header
st.markdown("""
<div class="main-header">
    <h1>🌐 Multilingual QnA Generator</h1>
    <p>Transform any document into intelligent Question-Answer pairs in English, Hindi & Marathi</p>
    <div class="feature-row">
        <span class="feature-badge">🤖 ADK Multi-Agent</span>
        <span class="feature-badge">📄 PDF / DOCX / TXT</span>
        <span class="feature-badge">🌍 3 Languages</span>
        <span class="feature-badge">📊 Excel Export</span>
        <span class="feature-badge">⚡ Gemini AI</span>
    </div>
</div>
""", unsafe_allow_html=True)

# Workflow steps
st.markdown("""
<div class="step-indicator">
    <div class="step-number">1</div>
    <div class="step-text">Upload a document (.pdf, .docx, or .txt)</div>
</div>
<div class="step-indicator">
    <div class="step-number">2</div>
    <div class="step-text">Configure QnA count and model in the sidebar</div>
</div>
<div class="step-indicator">
    <div class="step-number">3</div>
    <div class="step-text">Click "Generate" — ADK agents handle the rest!</div>
</div>
""", unsafe_allow_html=True)

st.markdown("")

# File uploader
uploaded_file = st.file_uploader(
    "📂 Upload Your Document",
    type=["pdf", "docx", "txt"],
    help="Upload a document in PDF, DOCX, or TXT format.",
)

# File info display
if uploaded_file:
    file_size = uploaded_file.size
    size_str = (
        f"{file_size / 1024:.1f} KB"
        if file_size < 1024 * 1024
        else f"{file_size / (1024 * 1024):.2f} MB"
    )
    st.markdown(
        f'<div class="stats-row">'
        f'<div class="stat-box"><div class="stat-value">📄</div>'
        f'<div class="stat-label">{uploaded_file.name}</div></div>'
        f'<div class="stat-box"><div class="stat-value">{size_str}</div>'
        f'<div class="stat-label">File Size</div></div>'
        f'<div class="stat-box"><div class="stat-value">{uploaded_file.name.split(".")[-1].upper()}</div>'
        f'<div class="stat-label">Format</div></div>'
        f'<div class="stat-box"><div class="stat-value">{num_pairs}</div>'
        f'<div class="stat-label">QnA Pairs</div></div>'
        f'</div>',
        unsafe_allow_html=True,
    )

# Generate button
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    generate_btn = st.button(
        "🚀 Generate Multilingual QnA",
        use_container_width=True,
        type="primary",
        disabled=not (uploaded_file and api_key),
    )

if not api_key:
    st.info("👈 Please configure your **Google API Key** in the sidebar.")


# ─── Pipeline Execution ─────────────────────────────────────────────

async def run_adk_pipeline(file_path: str, api_key: str, num_pairs: int, model: str, output_path: str):
    """Run the ADK multi-agent pipeline and yield events."""
    if api_key == "DEMO_MODE" or not api_key:
        api_key = "MOCK_KEY"
        model = "mock-model"
    os.environ["GOOGLE_API_KEY"] = api_key

    from google.adk.agents import LlmAgent, SequentialAgent
    from google.adk.runners import Runner
    from google.adk.sessions import InMemorySessionService
    from google.genai import types

    from multilingual_qna.tools.document_tools import parse_document, get_document_text
    from multilingual_qna.tools.qna_tools import save_qna_pairs, get_english_qna
    from multilingual_qna.tools.excel_tools import compile_excel_report
    from multilingual_qna.prompts import (
        DOCUMENT_PARSER_INSTRUCTION,
        QNA_GENERATOR_INSTRUCTION,
        HINDI_TRANSLATOR_INSTRUCTION,
        MARATHI_TRANSLATOR_INSTRUCTION,
        EXCEL_COMPILER_INSTRUCTION,
    )

    # Create agents
    document_parser = LlmAgent(
        name="document_parser",
        model=model,
        instruction=DOCUMENT_PARSER_INSTRUCTION,
        description="Parses documents and extracts text.",
        tools=[parse_document],
        output_key="parsed_document_output",
    )

    qna_generator = LlmAgent(
        name="qna_generator",
        model=model,
        instruction=QNA_GENERATOR_INSTRUCTION.format(num_pairs=num_pairs),
        description="Generates English QnA pairs from document text.",
        tools=[get_document_text, save_qna_pairs],
        output_key="english_qna_output",
    )

    hindi_translator = LlmAgent(
        name="hindi_translator",
        model=model,
        instruction=HINDI_TRANSLATOR_INSTRUCTION,
        description="Translates English QnA pairs to Hindi.",
        tools=[get_english_qna, save_qna_pairs],
        output_key="hindi_qna_output",
    )

    marathi_translator = LlmAgent(
        name="marathi_translator",
        model=model,
        instruction=MARATHI_TRANSLATOR_INSTRUCTION,
        description="Translates English QnA pairs to Marathi.",
        tools=[get_english_qna, save_qna_pairs],
        output_key="marathi_qna_output",
    )

    excel_compiler = LlmAgent(
        name="excel_compiler",
        model=model,
        instruction=EXCEL_COMPILER_INSTRUCTION.format(output_path=output_path),
        description="Compiles QnA pairs into Excel file.",
        tools=[compile_excel_report],
        output_key="excel_output_summary",
    )

    pipeline = SequentialAgent(
        name="multilingual_qna_pipeline",
        description="Full QnA generation pipeline.",
        sub_agents=[
            document_parser,
            qna_generator,
            hindi_translator,
            marathi_translator,
            excel_compiler,
        ],
    )

    # Run the pipeline
    session_service = InMemorySessionService()
    runner = Runner(
        agent=pipeline,
        app_name="multilingual_qna_app",
        session_service=session_service,
    )

    user_message = f"Process the document at: {file_path}"
    content = types.Content(
        role="user",
        parts=[types.Part.from_text(text=user_message)],
    )

    # Create session first
    await session_service.create_session(
        app_name="multilingual_qna_app",
        user_id="streamlit_user",
        session_id="streamlit_session",
    )

    async for event in runner.run_async(
        user_id="streamlit_user",
        session_id="streamlit_session",
        new_message=content,
    ):
        yield event



def extract_qna_from_events(events) -> dict:
    """Extract QnA JSON data from ADK pipeline events."""
    results = {"english": [], "hindi": [], "marathi": []}

    for event in events:
        # Check for agent responses that contain QnA JSON
        if not hasattr(event, 'content') or not event.content:
            continue
        if not hasattr(event.content, 'parts'):
            continue

        for part in event.content.parts:
            if not hasattr(part, 'text') or not part.text:
                continue
            text = part.text

            # Try to extract JSON arrays from the text
            json_matches = re.findall(r'\[\s*\{.*?\}\s*\]', text, re.DOTALL)
            for match in json_matches:
                try:
                    pairs = json.loads(match)
                    if isinstance(pairs, list) and len(pairs) > 0:
                        sample = pairs[0]
                        q = sample.get('question', sample.get('Question', ''))
                        a = sample.get('answer', sample.get('Answer', ''))
                        if q and a:
                            # Determine language by checking for Devanagari script
                            has_devanagari = bool(re.search(r'[\u0900-\u097F]', q))
                            if not has_devanagari and not results["english"]:
                                results["english"] = pairs
                            elif has_devanagari:
                                if not results["hindi"]:
                                    results["hindi"] = pairs
                                elif not results["marathi"]:
                                    results["marathi"] = pairs
                except (json.JSONDecodeError, Exception):
                    continue

    return results


# ─── Execution Logic ─────────────────────────────────────────────────

if uploaded_file and api_key and generate_btn:
    try:
        # Save uploaded file to temp location
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, uploaded_file.name)
        with open(temp_path, "wb") as f:
            uploaded_file.seek(0)
            f.write(uploaded_file.read())

        output_path = os.path.join(temp_dir, "QnA.xlsx")

        # Show agent pipeline progress
        st.markdown("### 🤖 ADK Agent Pipeline")

        agents_info = [
            ("📄 document_parser", "Extracting text from document..."),
            ("🧠 qna_generator", f"Generating {num_pairs} English QnA pairs..."),
            ("🇮🇳 hindi_translator", "Translating to Hindi (हिन्दी)..."),
            ("🇮🇳 marathi_translator", "Translating to Marathi (मराठी)..."),
            ("📊 excel_compiler", "Creating styled Excel output..."),
        ]

        progress_bar = st.progress(0, text="Initializing ADK agent pipeline...")

        # Run the pipeline
        events = []
        with st.spinner("🚀 Running ADK multi-agent pipeline..."):
            async def _consume():
                async for event in run_adk_pipeline(
                    file_path=temp_path,
                    api_key=api_key,
                    num_pairs=num_pairs,
                    model=model_name,
                    output_path=output_path,
                ):
                    events.append(event)
                    # Update progress bar dynamically based on the current agent executing
                    author = getattr(event, 'author', '')
                    if author == 'document_parser':
                        progress_bar.progress(0.2, text="Agent: 📄 document_parser — Parsing document text...")
                    elif author == 'qna_generator':
                        progress_bar.progress(0.4, text=f"Agent: 🧠 qna_generator — Generating {num_pairs} English pairs...")
                    elif author == 'hindi_translator':
                        progress_bar.progress(0.6, text="Agent: 🇮🇳 hindi_translator — Translating QnA to Hindi...")
                    elif author == 'marathi_translator':
                        progress_bar.progress(0.8, text="Agent: 🇮🇳 marathi_translator — Translating QnA to Marathi...")
                    elif author == 'excel_compiler':
                        progress_bar.progress(0.95, text="Agent: 📊 excel_compiler — Creating styled Excel output...")

            try:
                loop = asyncio.get_event_loop()
                if loop.is_running():
                    import nest_asyncio
                    nest_asyncio.apply()
                    loop.run_until_complete(_consume())
                else:
                    asyncio.run(_consume())
            except RuntimeError:
                asyncio.run(_consume())

        progress_bar.progress(1.0, text="✅ All agents completed successfully!")
        time.sleep(0.5)


        # Extract results from events
        results = extract_qna_from_events(events)

        # If results are empty, try to read from the Excel file directly
        if not any(results.values()) and os.path.exists(output_path):
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            for lang, sheet_name in [("english", "English"), ("hindi", "Hindi"), ("marathi", "Marathi")]:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    pairs = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:
                            pairs.append({"question": str(row[0]), "answer": str(row[1])})
                    results[lang] = pairs

        # Success banner
        st.markdown(
            '<div class="success-banner">'
            '✅ QnA Generation Complete — All 3 Languages Ready!'
            '</div>',
            unsafe_allow_html=True,
        )

        # Results in tabs
        tab_en, tab_hi, tab_mr = st.tabs(
            ["🇬🇧 English", "🇮🇳 Hindi (हिन्दी)", "🇮🇳 Marathi (मराठी)"]
        )

        for tab, lang_key, lang_label in [
            (tab_en, "english", "English"),
            (tab_hi, "hindi", "Hindi"),
            (tab_mr, "marathi", "Marathi"),
        ]:
            with tab:
                pairs = results.get(lang_key, [])
                st.markdown(f"### {lang_label} QnA Pairs ({len(pairs)} pairs)")
                for i, pair in enumerate(pairs, 1):
                    q = pair.get("question", pair.get("Question", ""))
                    a = pair.get("answer", pair.get("Answer", ""))
                    with st.container():
                        st.markdown(f"**Q{i}.** {q}")
                        st.markdown(f"**A{i}.** {a}")
                        st.markdown("---")

        # Excel download
        st.markdown("")
        st.markdown("### 📥 Download Excel Output")

        if os.path.exists(output_path):
            with open(output_path, "rb") as f:
                excel_data = f.read()
        else:
            # Generate Excel from results
            from multilingual_qna.tools.excel_tools import compile_excel_report
            english_json = json.dumps(results.get("english", []), ensure_ascii=False)
            hindi_json = json.dumps(results.get("hindi", []), ensure_ascii=False)
            marathi_json = json.dumps(results.get("marathi", []), ensure_ascii=False)
            compile_excel_report(english_json, hindi_json, marathi_json, output_path)
            with open(output_path, "rb") as f:
                excel_data = f.read()

        col_a, col_b, col_c = st.columns([1, 2, 1])
        with col_b:
            st.download_button(
                label="📥 Download QnA.xlsx",
                data=excel_data,
                file_name="QnA.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        st.markdown(
            '<div class="info-card">'
            '<h3>📊 Excel File Details</h3>'
            '<ul>'
            '<li><strong>File Name:</strong> QnA.xlsx</li>'
            '<li><strong>Sheet 1:</strong> English — QnA pairs in English</li>'
            '<li><strong>Sheet 2:</strong> Hindi — QnA pairs in हिन्दी</li>'
            '<li><strong>Sheet 3:</strong> Marathi — QnA pairs in मराठी</li>'
            '<li><strong>Columns:</strong> Questions, Answers</li>'
            '</ul>'
            '</div>',
            unsafe_allow_html=True,
        )

        # Show pipeline events (debugging)
        with st.expander("🔍 ADK Pipeline Events (Debug)", expanded=False):
            for i, event in enumerate(events):
                if hasattr(event, 'content') and event.content:
                    agent = getattr(event, 'author', 'unknown')
                    st.markdown(f"**Event {i+1}** — Agent: `{agent}`")
                    if hasattr(event.content, 'parts'):
                        for part in event.content.parts:
                            if hasattr(part, 'text') and part.text:
                                st.text(part.text[:500])
                    st.markdown("---")

        # Cleanup temp dir
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass

    except ImportError as e:
        st.error(
            f"❌ **Missing dependency**: {str(e)}\n\n"
            "Install required packages:\n"
            "```\npip install google-adk PyPDF2 python-docx openpyxl nest-asyncio\n```"
        )
    except Exception as e:
        st.error(f"❌ An error occurred: {str(e)}")
        st.markdown("**Possible fixes:**")
        st.markdown("- Verify your Google API Key is valid")
        st.markdown("- Ensure the document has readable text content")
        st.markdown("- Try a smaller document or fewer QnA pairs")
        with st.expander("🔧 Full Error Details"):
            st.exception(e)
